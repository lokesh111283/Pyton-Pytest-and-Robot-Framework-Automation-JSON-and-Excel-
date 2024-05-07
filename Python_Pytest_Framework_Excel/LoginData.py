import openpyxl

#Below code reads the data from excel and passes back to called metods in form of dictionary/JSON data
class LoginData:
    @staticmethod
    def getdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:\\html2\\logindata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]